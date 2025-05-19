#!/usr/bin/env python
# This script will take 2 sets of input vcf files (somaticseq output) i. Normals and ii. Samples
# Error formula is same as MAF as described in GenomeBiology,2019,20:50

import argparse
# import pysam
import sys, os, re
import pandas as pd
import numpy as np
from openpyxl import load_workbook

variants = []	# list of dictionaries
index = {}		# storing indices of keys

def parse_arguments():
	parser = argparse.ArgumentParser(description="Calculation of Error rate")
	parser.add_argument('--samples', action='append', nargs='*')	# positional argument
	parser.add_argument('--normals', action='append', nargs='*')
	return parser.parse_args()

def chromosome_parse(CHROM):
	if re.search('[a-zA-Z]', str(CHROM)):
		chrom = re.sub ("chr","", CHROM, flags = re.IGNORECASE)
		chrom = re.sub ("X","23", chrom, flags = re.IGNORECASE)
		chrom = re.sub ("Y","y", chrom, flags = re.IGNORECASE)
	return chrom

def extract_normals_error(file):
	# vcf_file = pysam.VariantFile(file)
	# for record in vcf_file:
	# 	print (record.contig, record.pos, record.ref, record.alts)
	# vcf_file.close()
	df = pd.read_excel(file, sheet_name=0, usecols=["Chr", "Start", "End", "Ref", "Alt", "REF_COUNT", "ALT_COUNT"])
	for ind, row in df.iterrows():
		chr = chromosome_parse(row[0])
		start = int(row[1])
		end = int(row[2])
		ref = str(row[3])
		alt = str(row[4])
		coverage = int(row[5]) + int(row[6])
		alt_count = int(row[6])
		key = (chr, start, end, ref, alt)
		if key in index:
			i = index[key]
			# variants[i].update({'coverage': (variants[i]['coverage'].append(coverage))})
			# variants[i].update({'alt_count': (variants[i]['alt_count'].append(alt_count))})
			variants[i]['coverage'].append(coverage)
			variants[i]['alt_count'].append(alt_count)
		else:
			coverage_list = [coverage]
			alt_count_list = [alt_count]
			new_variant = {'chr': chr, 'start': start, 'end': end, 'ref': ref, 'alt': alt, 'coverage': coverage_list, 'alt_count': alt_count_list}
			index[key] = len(variants)
			variants.append(new_variant)

def update_freq(samp_file):
	# Load the workbook and select the sheet
	workbook = load_workbook(samp_file)
	sheet = workbook.active  # or workbook['Sheet1'] if you know the sheet name
	# Define new headers and values to add
	new_headers = ['Background', 'Correctd. Ref', 'Correctd. Alt', 'Correctd. VAF %']
	# Find where to place new columns (after the VAF column)
	start_col = 13	
	# Inserting new columns 
	sheet.insert_cols(start_col, amount=len(new_headers))
	# Add new headers
	for i, header in enumerate(new_headers):
		sheet.cell(row=1, column=start_col + i, value=header)
	
	for row in sheet.iter_rows(min_row=2): 
		chr = chromosome_parse(row[0].value)
		start = row[1].value
		end = row[2].value
		ref = row[3].value
		alt = row[4].value
		ref_count = row[9].value
		alt_count = row[10].value
		samp_key = (chr, start, end, ref, alt)

		background = error_rate = correctd_ref = correctd_alt = correctd_vaf = 0
		if samp_key in index:
			ind = index[samp_key]
			# Calculating the error rate as mentioned in the paper GenomeBiology,2019,20:50
			error_rate = (variants[ind]['alt_count'] / variants[ind]['coverage'])
			# Setting the background to 3 x the error rate
			background = error_rate * 3
			# Correcting the ref and alt values by subtracting the background
			correctd_ref = ref_count - int((variants[ind]['coverage'] - variants[ind]['alt_count']) * background)
			correctd_alt = alt_count - int(variants[ind]['alt_count'] * background)
			correctd_vaf = (correctd_alt / (correctd_ref + correctd_alt)) * 100
		sheet.cell(row=row[0].row, column=start_col, value=background)
		sheet.cell(row=row[0].row, column=start_col + 1, value=correctd_ref)
		sheet.cell(row=row[0].row, column=start_col + 2, value=correctd_alt)
		sheet.cell(row=row[0].row, column=start_col + 3, value=correctd_vaf)

	modified_filename = samp_file.replace('.xlsx', '_mod.xlsx')
	workbook.save(modified_filename)

def main ():
	args = parse_arguments()
	# print(type (args.samples), type (args.normals))
	if args.normals:
		for normals in args.normals[0]:
			# print (normals)
			if os.path.getsize(normals) != 0:
				extract_normals_error(normals)	# make a list containing positions and their frequencies
		# updating the coverage and alt_count in the variants list to their median values
		for key, vals in enumerate(variants):
			median_coverage = np.median(np.array(vals['coverage']))
			median_alt = np.median(np.array(vals['alt_count']))
			variants[key].update({'coverage': median_coverage})
			variants[key].update({'alt_count': median_alt})

	if args.samples:
		for samples in args.samples[0]:
			if os.path.getsize(samples) != 0:
				update_freq(samples) 

if __name__ == '__main__':
	main()